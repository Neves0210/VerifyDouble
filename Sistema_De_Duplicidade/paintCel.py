import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

arquivo_excel = 'C:/Users/gabar/Downloads/Pasta1.xlsx'

df = pd.read_excel(arquivo_excel)

coluna_alvo = 'Unnamed: 1'

duplicados = df[df.duplicated(coluna_alvo, keep=False)]

wb = load_workbook(arquivo_excel)
ws = wb.active

red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

def is_cell_red(cell):
    return cell.fill.start_color.index == "FFFF0000" and cell.fill.end_color.index == "FFFF0000"

pintados = set()

for index, row in duplicados.iterrows():
    cell = ws.cell(row=index + 2, column=df.columns.get_loc(coluna_alvo) + 1)
    valor = row[coluna_alvo]

    if valor in pintados:
        continue

    if not is_cell_red(cell):
        cell.fill = red_fill  
        pintados.add(valor) 

wb.save(arquivo_excel)
